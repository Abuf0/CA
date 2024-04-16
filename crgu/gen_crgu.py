import openpyxl
class clock:
    def __init__(self,src,dest,div,ckgt,mux_sel,mux_source,to_module):
        self.src = src
        self.dest = dest
        self.div = div
        self.ckgt = ckgt
        self.mux = max
        self.mux_sel = mux_sel
        self.mux_source = mux_source
        self.to_module = to_module
    def print_info(self):
        print("clock path : "+self.src+" --> "+self.dest)
        print("dest : ",self.dest)
        print("src  : ",self.src)
        print("div  : ",self.div)
        print("ckgt : ",self.ckgt)
        print("mux_sel : ",self.mux_sel)
        print("mux_source : ",self.mux_source)
        print("to_module : ",self.to_module)

    '''
    def connect(self):
        if(self.div==1):
            if(self.ckgt==None):
                print("\033[93m"+self.src + " --> SCAN_MUX(scan_clk) --> " + self.dest+"\033[0m")
            else:
                print("\033[93m"+self.src +" --> SCAN_MUX(scan_clk) --> ICG("+ self.ckgt +") --> " + self.dest+"\033[0m")
        elif(self.div>1):
            if(self.ckgt==None):
                print("\033[93m"+self.src +" --> SCAN_MUX(scan_clk) --> DIV("+ str(self.div) +") --> SCAN_MUX(scan_clk) --> " + self.dest+"\033[0m")
            else:
                print("\033[93m"+self.src +" --> SCAN_MUX(scan_clk) --> DIV("+ str(self.div) +") --> ICG("+ self.ckgt +") --> " + self.dest+"\033[0m")
    '''
    def connect(self):
        if(self.mux_sel==None):
            if(self.div==1):
                if(self.ckgt==None):
                    print("\033[93m"+self.src + " --> SCAN_MUX(scan_clk) --> " + self.dest+"\033[0m")
                else:
                    print("\033[93m"+self.src +" --> SCAN_MUX(scan_clk) --> ICG("+ self.ckgt +") --> " + self.dest+"\033[0m")
            elif(self.div>1):
                if(self.ckgt==None):
                    print("\033[93m"+self.src +" --> SCAN_MUX(scan_clk) --> DIV("+ str(self.div) +") --> SCAN_MUX(scan_clk) --> " + self.dest+"\033[0m")
                else:
                    print("\033[93m"+self.src +" --> SCAN_MUX(scan_clk) --> DIV("+ str(self.div) +") --> ICG("+ self.ckgt +") --> " + self.dest+"\033[0m")


                

class reset:
    def __init__(self,dest_rstn,sync_clk,ls_grstn,ls_lrstn):
        self.dest_rstn = dest_rstn
        self.sync_clk = sync_clk
        self.ls_grstn = ls_grstn
        self.ls_lrstn = ls_lrstn
    def print_info(self):
        print("reset : ",self.dest_rstn)
        print("sync_clk : ",self.sync_clk)
        print("global rstn  : ",self.ls_grstn)
        print("local rstn  : ",self.ls_lrstn)
    def connect(self):
        #print("//////////////////////////////////////////////")
        flag = 0
        for grstn in self.ls_grstn:
            if(flag==0):
                print("\033[93m"+grstn,end='')
            else:
                print(' & '+grstn,end='')
            flag = 1
        if(len(self.ls_lrstn)>0):
            print('  ->',end='')
            for lrstn in self.ls_lrstn:
                print(' & '+lrstn,end=' ')
        print('  ==>  SCAN_MUX(scan_rstn) --> SYNC('+self.sync_clk+') --> SCAN_MUX(scan_rstn)',end='')
        print('  ==>  ' + self.dest_rstn+"\033[0m")
        #print("//////////////////////////////////////////////")

# openpyxl.load_workbook(需要打开的excel文件路径)
wb = openpyxl.load_workbook('crgu.xlsx')
# 获取所有表的表名
sheets_names = wb.sheetnames
#print(sheets_names)     # 结果: ['表1', '表2']

# 根据表名获取工作簿中指定的表
sheet_clock = wb['clock']
#print(sheet_clock)           # 结果：<Worksheet "表2">

sheet_reset = wb['reset']
#print(sheet_reset)


### clock
print("\033[91m GENERATE CLOCK PATH \033[0m")
row_num = sheet_clock.max_row     # 获取当前表中最大的行数
col_num = sheet_clock.max_column
#print("sheet clock row & column:")
#print(row_num)
#print(col_num)
clock_num = col_num-1
objects = []
ls_ckgt = []
ls_mux_source = []
ls_to_module = []
#print(clock_num)
## 实例化类对象
for col in range(2,col_num+1):
    ls_ckgt = []
    ls_mux_source = []
    ls_to_module = []
    dest = sheet_clock.cell(1, col).value
    src = sheet_clock.cell(3, col).value
    div = sheet_clock.cell(5, col).value
    #ckgt = sheet_clock.cell(6, col).value
    for i in range(6,row_num+1):
        if(sheet_clock.cell(i, 1).value == 'mux_sel'):
            mux_sel_row = i
            break
        ls_ckgt.append(sheet_clock.cell(i, col).value)
    mux_sel = sheet_clock.cell(mux_sel_row, col).value
    for i in range(mux_sel_row+1,row_num+1):
        if(sheet_clock.cell(i, 1).value == 'to_module'):
            to_module_row = i
            break
        ls_mux_source.append(sheet_clock.cell(i, col).value)
    for i in range(to_module_row,row_num+1):
        if(sheet_clock.cell(i, col).value == None):
            break    
        ls_to_module.append(sheet_clock.cell(i, col).value)
    obj = clock(src,dest,div,ls_ckgt,mux_sel,ls_mux_source,ls_to_module)
    objects.append(obj)
## 操作类
print("----------------------------------------------------------------")
for obj in objects:
    obj.print_info()
    obj.connect()
    print("----------------------------------------------------------------")

### reset
print("\033[91m GENERATE RESET PATH \033[0m")
row_num = sheet_reset.max_row     # 获取当前表中最大的行数
col_num = sheet_reset.max_column
#print("sheet reset row & column:")
#print(row_num)
#print(col_num)
clock_num = col_num-1

objects = []

for row in range(1,row_num+1):
    if(sheet_reset.cell(row, 1).value=='global rstn'):
        grstn_id0 = row
    if(sheet_reset.cell(row, 1).value=='local rstn'):
        grstn_id1 = row-1
        lrstn_id0 = row
    lrstn_id1 = row_num
#print(grstn_id0)
#print(grstn_id1)
#print(lrstn_id0)
#print(lrstn_id1)

## 实例化类对象
for col in range(2,col_num+1):
    dest_rstn = sheet_reset.cell(1, col).value
    sync_clk = sheet_reset.cell(2, col).value
    ls_grstn = []
    ls_lrstn = []
    for row in range(grstn_id0,grstn_id1+1):
        if(sheet_reset.cell(row, col).value != None):
            ls_grstn.append(sheet_reset.cell(row, col).value)
    for row in range(lrstn_id0,lrstn_id1+1):
        if(sheet_reset.cell(row, col).value != None):
            ls_lrstn.append(sheet_reset.cell(row, col).value)
    #print(ls_grstn)
    #print(ls_lrstn)
    obj = reset(dest_rstn,sync_clk,ls_grstn,ls_lrstn)
    objects.append(obj)

## 操作类
print("----------------------------------------------------------------")
for obj in objects:
    #obj.print_info()
    #obj.connect()
    print("----------------------------------------------------------------")